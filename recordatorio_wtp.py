from openpyxl import load_workbook #Importamos librería que deja cargar excel (solo .xlsx)
import time
import os #libreria para cerrar ventanas y .exe
from selenium import webdriver #Importamos selenium
from selenium.webdriver.common.keys import Keys  #Para realizar acciones por teclado
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait #Para hacer las esperas y validar condiciones
from selenium.webdriver.support import expected_conditions as EC #Condiciones, modulo de selenium
import pyttsx3
import pyautogui as pg, webbrowser as web
from email.mime import text
import smtplib
from email import encoders ##para codificar los elementos a adjuntar
from email.mime.base import MIMEBase 
from email.mime.text import MIMEText ##mime: (mensaje multimedia)
from email.mime.multipart import MIMEMultipart
from typing import FrozenSet #Ya vienen preinstalados (mensaje multiparte (adjunto, asunto etc))


def whatsapp():
    print("WhatsApp")
    #Webdriver
    driver = webdriver.Chrome(executable_path="C:/Users/luise/Downloads/chromedriver.exe") #Declaramos la función abrir navegador con webdriver
    driver.get("https://web.whatsapp.com/")
    print("Tiene 30 segundos para validar su sesión...")
    time.sleep(30)
    #Excel
    excel = "./Base de datos.xlsx"
    wb = load_workbook(excel)
    hojas= wb.get_sheet_names() #Permite saber cuantas hojas tiene el excel 
    print(hojas, "Hojas en el excel")
    hoja1 = wb.get_sheet_by_name('Hoja1')
    #mensaje = hoja1[f'B1'] #Lee el mensaje a escribir
    #print (mensaje.value, "Es el Mensaje a enviar")
    #ruta_adjunto = hoja2[f'B2'] #Lee la ruta del archivo a adjuntar
    #print (ruta_adjunto.value, "Es la ruta del adjunto")

    #Preparamos para el ciclo de Whatsapp
    contador_excel = 2
    ultima_fila = hoja1.max_row
    ultima_fila = ultima_fila + 1
    print(ultima_fila, "ultima fila")

    for i in range((contador_excel), (ultima_fila)):
        telefono = hoja1[f'E{contador_excel}']
        print(telefono.value)
        if str(telefono.value) == 'None':
            print(telefono.value)
        else:
            
            driver.get("https://web.whatsapp.com/send?phone=+57" + str(telefono.value)) #Llamamos la función abrir navegador
            time.sleep(5)
            try:
                chat = WebDriverWait(driver, 8).until(EC.element_to_be_clickable((By.XPATH,"/html/body/div/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div[2]/div/div[1]/div/div[2]")))
                print(contador_excel, "Chat encontrado")
                
                mensaje_g = hoja1[f'G{contador_excel}']
                print (mensaje_g.value, "Es el saludo")
                fecha_h = hoja1[f'H{contador_excel}']
                print (fecha_h.value, "Es la fecha")
                hora_i = hoja1[f'I{contador_excel}']
                print(hora_i.value, "Es la hora")
                direccion_j = hoja1[f'J{contador_excel}']
                print(direccion_j.value, "es el lugar")
                profesional_k = hoja1[f'K{contador_excel}']
                print(profesional_k.value, "Medico")
                recomendaciones_l = hoja1[f'L{contador_excel}']
                print(recomendaciones_l.value, "son las recomendaciones")
                #Concatenamos mensaje
                mensaje = str(mensaje_g.value +" el día " +  fecha_h.value +" A las "+ hora_i.value + " En la " + direccion_j.value +" Con el Médico "+ profesional_k.value +" le recordamos "+ recomendaciones_l.value)
                print(mensaje)

                #Escribimos y enviamos mensaje de texto
                driver.find_element_by_xpath("//*[@id='main']/footer/div[1]/div[2]/div/div[1]/div/div[2]").send_keys(mensaje) #Escribimos el mensaje!
                driver.find_element_by_xpath("/html/body/div/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div[2]/div/div[2]/button").click() #click en enviar


                #Convertimos mensaje a mp3
                engine = pyttsx3.init()
                engine.setProperty("rate",150)
                telefono = str(telefono.value)
                audio_guardado = "Audio"+(telefono)+".mp3"
                engine.save_to_file(mensaje, audio_guardado)
                #engine.say(mensaje) #reproduce
                engine.runAndWait()

                #Adjuntamos y enviamos el audio
                driver.find_element_by_xpath("/html/body/div/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div[1]/div[2]/div/div/span").click()#click en adjunto
                time.sleep(2)
                driver.find_element_by_xpath("/html/body/div/div[1]/div[1]/div[4]/div[1]/footer/div[1]/div[1]/div[2]/div/span/div[1]/div/ul/li[1]/button/span").click() #imgicon
                time.sleep(1)
        
                audio_guardado = "\Audio"+(telefono)+".mp3"
                ruta_audio = (r"C:\Users\luise" + str(audio_guardado))
                pg.write(ruta_audio)
                time.sleep(1)
                pg.press('enter') #damos enter en cuadro de dialogo 
                time.sleep(1)
                driver.find_element_by_xpath("/html/body/div/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/span/div/div/span").click()#enviar adj

                time.sleep(8) #esperamos que salga el mensaje y cargue el nuevo número



            except:
                time.sleep(3)
                driver.find_element_by_xpath("/html/body/div/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div[2]/div/div/div").click()
                print(contador_excel, "Chat no encontrado")
            #leemos las base de datos
        
        contador_excel = contador_excel + 1
            #
            # web.open("https://web.whatsapp.com/send?phone=+" + str(telefono.value))
            #Celdas de excel de hora y fecha deben estar en formato texto

def correo():
    print("Correo")
    servidor = smtplib.SMTP('smtp.gmail.com', 587) ##conectamos al servidor
    servidor.starttls()

    #origen = "incapacidades.100digital@gmail.com"



    #Excel
    excel = "./Base de datos.xlsx"
    wb = load_workbook(excel)
    hojas= wb.get_sheet_names() #Permite saber cuantas hojas tiene el excel 
    print(hojas, "Hojas en el excel")
    hoja1 = wb.get_sheet_by_name('Hoja1')
    hoja2 = wb.get_sheet_by_name('Hoja2')

    #Extraemos El asunto Del Correo
    asunto_extraido = hoja2[f'B3'] #extraemos el asunto de excel
    asunto = str(asunto_extraido.value) #Damos formato al asunto
    #Extraemos el Correo de origen
    correo_extraido = hoja2[f'B1'] #Leemos el correo de excel
    origen = str(correo_extraido.value) #Damos formato al correo extraído y lo asignamos a la variable del multiparte
    #Extraémos la contraseña
    pass_extraida = hoja2[f'B2'] #Leemos la contraseña del excel
    pass_origen = str(pass_extraida.value) #Damos formato al pass extraido del excel

    #Extraemos el nombre del adjunto
    archivo_extraído = hoja2[f'B4']
    archivo = str(archivo_extraído.value) #Lee el nombre o la ruta del adjunto

    servidor.login(origen, pass_origen) #pasamos los datos del correo de origen para iniciar sesión

    contador_excel = 2 
    ultima_fila = hoja1.max_row
    ultima_fila = ultima_fila + 1
    print(ultima_fila, "ultima fila")

    for i in range ((contador_excel), (ultima_fila)):
        destino = hoja1[f'F{contador_excel}']
        print (destino.value)
        if str(destino.value) == 'No hay email' or str(destino.value) == 'None':
            print(destino.value, 'No enviado', contador_excel)
        else:
            #Extraemos Cuerpo Del Mensaje
            cuerpom = hoja1[f'G{contador_excel}'] #lee el mensaje del excel
            cuerpo_mensaje = str(cuerpom.value) #pasamos a string y leemos el text
            ##Creamos la estructura del mensaje:
            mensaje = MIMEMultipart("alternative") #el correo será un mensaje multiparte de tipo estándar
            mensaje ["Subject"] = asunto #Pasamos la parte del asunto...
            mensaje ["From"] = origen #pasamos el correo de origen
            mensaje ["To"] = str(destino.value) #pasamos el correo de destino


            #Insertamos el html:
            html = f"""

            <html>
            <head>
                <meta charset="utf-8">
                <title>correo-robot</title>
            </head>
            <body style="background-color: white;">
            <header style = "border: 7px solid rgba(141, 149, 156, 0.637);"">
                <h1 style=" margin: 20px auto; text-align: center; color: rgb(80, 80, 43);">Hola, {str(destino.value)} Queremos Recordar Tu Próxima Cita</h1>
            </header>
            <section>
                <h2 style="margin: 10px auto; color: rgb(150, 143, 56); text-align: center; ">Clínica Marly</h2>
            </section>
            <section>
                <p style="text-align: justify center ;">{cuerpo_mensaje}</p>
                <a style="background-color: rgb(64, 88, 224);
                            display: block;
                            margin-left: auto;
                            margin-right: auto;
                            color: white;
                            padding: 10px;
                            margin: 10px;
                            font-size: 20px;
                            font-family:Cambria, Cochin, Georgia, Times, 'Times New Roman', serif;
                            text-decoration: none;
                            text-transform: uppercase;
                            font-weight: bold;
                            border-radius: 5px;
                            text-align: center;
                            
                            " href="https://marly.com.co/">Visítanos</a>
            </section>
            </body>
            </html>

            """

            parte_html = MIMEText(html, "html") #indicamos que es formato html

            mensaje.attach(parte_html) #adjuntamos a mensaje la parte html

            #anexamos el adjunto desde aquí
            #archivo = "instalar selenium.txt" #entre "" va la ruta donde está el archivo, acá queda así porque está en la misma carpeta del codigo

            with open(archivo, "rb") as adjunto: ##lea r como bytes b el archivo
                contenido_adjunto = MIMEBase("application","octet-stream") #Que se interprete como una aplicación (xlsx)
                contenido_adjunto.set_payload(adjunto.read())

            encoders.encode_base64(contenido_adjunto) #codificamos el adjunto

            contenido_adjunto.add_header(
                "Content-Disposition",
                f"attachment; filename= {archivo}",
            )

            mensaje.attach(contenido_adjunto) #añadimos al mensaje multiparte el adjunto
            mensaje_final = mensaje.as_string() #empaquetamos todo el correo

            #enviamos formalmente el correo
            servidor.sendmail(origen, str(destino.value), mensaje_final) #origen, destino y mensaje
            print('Correo', {contador_excel}, 'enviado')

        contador_excel = contador_excel + 1

                    


    #cerramos el servidor
    servidor.quit()

    print('Correo enviado')


menu = """
Bienvenid@, soy tu asistente de recordatorios de citas médicas 🏥🤖
¿Que tipo de recordatorio deseas enviar?

1- Mensajes y audios de WhatsApp 📲
2- Mensajes por Email 📧

Elige una opción: """


opcion = input(menu)
if opcion == '1':
    whatsapp()
elif opcion == '2':
    correo()
else:
    print(opcion +',no es una opción del menú, ''Por favor ingrese una opción correcta')